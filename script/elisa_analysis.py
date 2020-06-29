'''
title: elisa_analysis.py
author: Michael Korenkov
date: 2020/06/29
'''

import numpy as np
import pandas as pd
import itertools as it
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt


def load_data(file, n):
    """
    Loads OD values from excel file into lists
    """
    workbook = load_workbook(filename=file)
    data, data1, data2, data3, data4 = [], [], [], [], []
    for sheets in workbook.sheetnames:
        data.append(pd.read_excel(file, sheet_name=sheets, usecols="A:M", index_col=0, skiprows=10, nrows=8))
    if n == 1:
        data1 = data
        return data1
    elif n == 2:
        data1 = data[0::n]
        data2 = data[1::n]
        return data1, data2
    elif n == 3:
        data1 = data[0::n]
        data2 = data[1::n]
        data3 = data[2::n]
        return data1, data2, data3
    elif n == 4:
        data1 = data[0::n]
        data2 = data[1::n]
        data3 = data[2::n]
        data4 = data[3::n]
        return data1, data2, data3, data4
    return print(str(file) + " has been loaded")


def get_standard(array, i):
    """
    Gets the values for the antibody used as a standard on the plate
    and saves it into a new array while deleting itsself from the old array
    """
    axis = int(input("Enter the axis of the standard (1=1 OR 2=H): "))
    standard_array = []
    value_array = []
    for x in range(i):
        standard_array.append([])
        value_array.append([])
        for n in range(len(array[x])):
            if axis == 1:
                standard_array[x].append(array[x][n].iloc[0:8, 0].to_numpy())
                value_array[x].append(array[x][n].drop(1, axis=1))
            elif axis == 2:
                standard_array[x].append(array[x][n].iloc[[0]].to_numpy())
                value_array[x].append(array[x][n].drop(["H"]))
    return value_array, standard_array


def analysis(standard, values, plates):
    """
    Chooses which timepoints to plot by comparing the standard_array and
    finding the clostest partner
    """
    all_means = []
    standard_dict = {}  # all standard value computing (mean of sum of means) saved here
    export_standard_dict = {}  # for easy returnability of standard dictionary and keys
    value_dict = {}  # for easy returnability of value dictionary and keys
    for x in range(plates):
        for n in range(len(standard[x])):
            all_means.append(standard[x][n].mean())
    
    mean_standard = [list(i) for i in zip(*[all_means[i:i+plates] for i in range(0, len(all_means), plates)])]

    for x in range(plates):  # Section checks how many input platesband calculates the mean of the mean of the standard 
        if plates == 1:
            standard_dict[x] = mean_standard[x]
        if plates == 2:
            if x == plates - 1:  # for loop breakes before last loop
                break
            else:
                for n in range(len(mean_standard[x])):
                    for i in range(len(mean_standard[x])):
                        standard_dict[(x, n), (x+1, i)] =  (np.array(mean_standard[x][n]) + np.array(mean_standard[x+1][i]))/plates
        if plates == 3:
            if x == plates - 1:  # for loop breakes before last loop
                break
            else:
                for n in range(len(mean_standard[x])):
                    for i in range(len(mean_standard[x])):
                        standard_dict[(x, n), (x+1, i), (x+2, i)] =  (np.array(mean_standard[x][n]) + np.array(mean_standard[x+1][i]) + np.array(mean_standard[x+2][i]))/plates
        if plates == 4:
            if x == plates - 1:  # for loop breakes before last loop
                break
            else:
                for n in range(len(mean_standard[x])):
                    for i in range(len(mean_standard[x])):
                        standard_dict[(x, n), (x+1, i), (x+2,i), (x+3, i)] =  (np.array(mean_standard[x][n]) + np.array(mean_standard[x+1][i]) + np.array(mean_standard[x+2][i]) + np.array(mean_standard[x+3][i]))/plates

    for x in range(plates):  # Section to convert value array into a dictionary for easy return
        for n in range(len(values[x])):
            value_dict[(x, n)] = values[x][n]
    
    for x in range(plates):  # Section to convert standard array into a dictionary for easy return
        for n in range(len(standard[x])):
            export_standard_dict[(x, n)] = standard[x][n]

    min_key_pair = min(standard_dict, key=standard_dict.get)  # computes the key for the lowest mean pair
    key_a, key_b = min_key_pair
    print(key_a, key_b)
    
    if plates == 1:
        key_a = min_key_pair
        return export_standard_dict[key_a], value_dict[key_a]
    if plates == 2:
        key_a, key_b = min_key_pair
        return export_standard_dict[key_a], value_dict[key_a], export_standard_dict[key_b], value_dict[key_b]
    if plates == 3:
        key_a, key_b, key_c = min_key_pair
        return export_standard_dict[key_a], value_dict[key_a], export_standard_dict[key_b], value_dict[key_b], export_standard_dict[key_c], value_dict[key_c]
    if plates == 4:
        key_a, key_b, key_c, key_d = min_key_pair
        return export_standard_dict[key_a], value_dict[key_a], export_standard_dict[key_b], value_dict[key_b], export_standard_dict[key_c], value_dict[key_c], export_standard_dict[key_d],value_dict[key_d]


def write_output(filename, standard_array, value_array):
    """
    Writes the values of the choosen standard_array into a excel file for further analysis
    """
    workbook = Workbook()
    sheet = workbook.active
    for value in dataframe_to_rows(standard_array[0][0].to_frame()):
        sheet.append(value)
    workbook.save(filename=str(filename + ".xlsx"))


def create_plot(plot_data):
    """
    Plots ELISA values with x=concentration and y=OD(405nm)
    """
    plt.plot(np.logspace(0.001, 1, 7), plot_data[0].iloc[0:7, 0])
    plt.xscale("log")
    plt.xlabel("concentration")
    plt.ylabel("OD 405nm")
    plt.show()


excel_file = input("Please enter the name of the excel file: ")
plates = int(input("Please enter the number of plates: "))
data = list(load_data(excel_file, plates))
value_array, standard_array = get_standard(data, plates)
number_of_sheet = analysis(standard_array, value_array, plates)
print(number_of_sheet)
#print(type(standard_array[0][0]))
#write_output(input("Filename? "), standard_array, value_array)
